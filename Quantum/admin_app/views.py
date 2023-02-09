from django.shortcuts import render,redirect
from user.models import users
from django.contrib import auth
from django.contrib import messages
from django.views.decorators.cache import cache_control,never_cache
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.views.generic import TemplateView,View
from.models import *
from orders.models import Order
from vendor.models import Product
import io
from django.http import FileResponse
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from datetime import datetime
from django.db.models import Count, DateTimeField
from django.db.models.functions import Trunc
from django.contrib.auth.decorators import user_passes_test



def admin_check(user):
    return user.is_superadmin



# Create your views here.
@never_cache
@login_required(login_url='/admin')
@user_passes_test(admin_check)
def admin_pannel(request):
   
  order = Order.objects.all()
  cancelled_order = Order.objects.filter(status = 'Cancelled')
  vendor = users.objects.filter(Q(is_staff = True) and Q(is_admin = False))
  user = users.objects.filter(is_staff = False)
  vendor_count = vendor.count()
  user_count = user.count()




  
  if order:

    total_amount = 0
    count = 0
    for i in order:
        total_amount += i.amount
        count += 1

    total_cancelled_orders = 0

    for i in cancelled_order:
        total_cancelled_orders+=1
    # total orders
    date_to_total_orders = {}
    date_of_order = []

    


    orders = order
    annotated_orders = orders.annotate(date=Trunc('order_date', 'day', output_field=DateTimeField()))
    final_orders = annotated_orders.annotate(count=Count('id')).values('date','count')

    for i in final_orders:
      date = i['date'].strftime("%Y-%m-%d")
      if date in date_to_total_orders:
        date_to_total_orders[date] += i['count']
      else:
        date_to_total_orders[date] = i['count']
        date_of_order.append(date)


    
    #cancelled orders

    date_to_cancelled_orders = {}
    date_of_cancelled_order = []

    


    orders = cancelled_order.all()
    annotated_orders = orders.annotate(date=Trunc('order_date', 'day', output_field=DateTimeField()))
    final_orders = annotated_orders.annotate(count=Count('id')).values('date','count')

    for i in final_orders:
      date = i['date'].strftime("%Y-%m-%d")
      if date in date_to_cancelled_orders:
        date_to_cancelled_orders[date] += i['count']
      else:
        date_to_cancelled_orders[date] = i['count']
        date_of_cancelled_order.append(date)

    print(date_of_cancelled_order)
    current_year =  datetime.now().year

    context = {'date_to_total_orders':date_to_total_orders,
               'date_of_order':date_of_order,
               'current_year':current_year,
               'date_of_cancelled_order':date_of_cancelled_order,
               'date_to_cancelled_orders':date_to_cancelled_orders,
               'total_revenue':total_amount,
               'total_order':count,
               'total_cancelled_orders':total_cancelled_orders,
               'user_count':user_count,
               'vendor_count':vendor_count}
    return render( request,'admin_pannel.html',context)
  else:
    return render(request, 'admin_pannel.html', {})





  

#----------------------------------------------------------------- user management------------------------------------------------------------------------------------------------------
@never_cache
@login_required(login_url='/admin')
@user_passes_test(admin_check)
def user_management(request):
    user = {
        'user': users.objects.filter(is_staff = False)
    }
    
    return render(request,'user_management.html',user)



@login_required(login_url='/admin')
@user_passes_test(admin_check)
def user_delete(request,id):
    usr = users.objects.filter(id = id).delete()
    return redirect('admin_users')

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


# ------------------------------------------------------------- vendor management--------------------------------------------------------------------------------------------------------
@never_cache
@login_required(login_url='/admin')
@user_passes_test(admin_check)
def vendor_management(request):
    vendor ={
        'vendor': users.objects.filter( Q(is_staff = True) & Q(is_superadmin = False))
    }

    return render(request,'vendor_management.html',vendor)



@login_required(login_url='/admin')
@user_passes_test(admin_check)
def vendor_delete(request,id):
    users.objects.filter(id = id).delete()
    return redirect('admin_vendors')




@login_required(login_url='/admin')
@user_passes_test(admin_check)
def block(request,id):
    users.objects.filter(id = id).update(is_active = False) 
    messages.success(request,'blocked successfuly')

    if users.objects.filter(id = id).filter(is_staff = True):
        return redirect('admin_vendors')
    else:
        return redirect('admin_users')
    


@login_required(login_url='/admin')
@user_passes_test(admin_check)
def unblock(request,id):
    users.objects.filter(id = id).update(is_active = True)
    messages.success(request,'unblocked successfuly')

    if users.objects.filter(id = id).filter(is_staff = True):
        return redirect('admin_vendors')
    else:
        return redirect('admin_users')
    


    

#------------------------------------------------------------------------------------------Category and Brand ------------------------------------------------------------------------------------------------
# ===================================Categories=======================================================
@never_cache
@login_required(login_url='/admin')
@user_passes_test(admin_check)
def category_management(request):

   context =  { 'category':Category.objects.all(),

     }
  
   return render(request,'category_management.html',context)



@login_required(login_url='/admin')
@user_passes_test(admin_check)
def add_category(request):

    if request.method == 'POST':
        try:
          category_logo = request.FILES['logo']
          category_name = request.POST['category_name']
          commission = request.POST['commission']
        except:
            messages.warning(request,'Image  can\'t be empty')
            return redirect('admin_category')
            
        if   category_name == '':
            messages.warning(request,'Category name cant be empty')
            return redirect('admin_category')

        if Category.objects.filter(category_name__icontains = category_name).exists():
               messages.warning(request,'Category already exists')
               return redirect('add_category')
    try:
        category = Category(
              category_image = category_logo,
              category_name = category_name,
              commission = commission
              )

        category.save()

        messages.success(request,'Category added successfully.')
        return redirect('admin_category')
    except:
        messages.warning(request,'Check the details given by you and try agin with proper values.')
        print('check 6')
        return redirect('admin_category')




@login_required(login_url='/admin')
@user_passes_test(admin_check)
def edit_category(request,id):
    if request.method == 'POST':

        try:
          category_logo = request.FILES['logo']
          category_name = request.POST['category_name']
          commission = request.POST['commission']

        except:
            messages.warning(request,'Image or name can\'t be empty')
            return redirect('admin_category')


        if   category_name == '':
            messages.warning(request,'Category name cant be empty')
            return redirect('admin_category')   

        
        try:
         category = Category(
                id=id,
                category_image = category_logo,
                category_name = category_name,
                commission = commission,
                date_added = datetime.now(),
                last_update = datetime.now()
            )

         category.save()


         messages.success(request,'Category Updated successfully')
         return redirect('admin_category')
        except Exception as e:
            print(e)
            messages.warning(request,'check your credentials and try again')
            return redirect('admin_category')






@login_required(login_url='/admin')
@user_passes_test(admin_check)
def delete_category(request,id):
    
           Category.objects.filter(id=id).delete()
           messages.success(request,'Category deleted successfully')
           return redirect('admin_category')



# =====================================Brands==================================================
@never_cache
@login_required(login_url='/admin')
@user_passes_test(admin_check)
def brand_management(request):
    context = {
        'brand':Brand.objects.all()
    }
    return render(request,'Brands.html',context)


@login_required(login_url='/admin')
@user_passes_test(admin_check)
def add_brand(request):
    if request.method == 'POST':

        try:
          brand_logo = request.FILES['logo']
          brand_name = request.POST['brand_name']

        except:
            messages.warning(request,'Image can\'t be empty')
            return redirect('admin_brand')


        if   brand_name == '':
            messages.warning(request,'Brand name cant be empty')
            return redirect('admin_brand')   

        if Brand.objects.filter(brand_name__icontains = brand_name).exists():
            messages.warning(request,'Brand already exists')
            return redirect('admin_brand')
        else:

            brand = Brand(
                brand_logo = brand_logo,
                brand_name = brand_name
            )

            brand.save()


            messages.success(request,'Brand added successfully')
            return redirect('admin_brand')





@login_required(login_url='/admin')
@user_passes_test(admin_check)
def edit_brand(request,id):
    if request.method == 'POST':

        try:
          brand_logo = request.FILES['logo']
          brand_name = request.POST['brand_name']

        except:
            messages.warning(request,'Image or name can\'t be empty')
            return redirect('admin_brand')


        if   brand_name == '':
            messages.warning(request,'Brand name cant be empty')
            return redirect('admin_brand')   

        else:

            brand = Brand(
                id=id,
                brand_logo = brand_logo,
                brand_name = brand_name,
                date_added = datetime.now(),
                last_update = datetime.now()
            )

            brand.save()


            messages.success(request,'Brand Updated successfully')
            return redirect('admin_brand')





@login_required(login_url='/admin')
@user_passes_test(admin_check)
def delete_brand(request,id):
        Brand.objects.filter(id=id).delete()
        messages.success(request,'Brand deleted successfully')
        return redirect('admin_brand')


# ========================================================================Banner================================================================================================================
@never_cache
@login_required(login_url='/admin')
@user_passes_test(admin_check)
def banner(request):
    banner = {
        'banner':Banner.objects.all()
    }
    print(banner,'jhgfds')
    return render(request,'banner.html',banner)



@login_required(login_url='/admin')
@user_passes_test(admin_check)
def add_banner(request):
    if request.method == 'POST':

        try:
            banner_image = request.FILES['bannerimage']
            banner_title = request.POST['bannertitle']
            banner_description = request.POST['bannerdescription']
            url = request.POST['url']
        except:
            messages.warning(request,'please fill  the columns')

        if Banner.objects.filter(banner_title = banner_title):
            messages.warning(request,'A banner with the title already exists.')
            return redirect('admin_banners')

            
        try:
                banner = Banner(
                    banner_image = banner_image,
                    banner_title = banner_title,
                    banner_description = banner_description,
                    url = url
                )

                banner.save()
                return redirect("admin_banners")
            
        except:
                messages.warning(request,'Something went Wrong')
                return redirect('admin_banners')
            


        
    return redirect(request,'banner.html')




@login_required(login_url='/admin')
@user_passes_test(admin_check)
def edit_banner(request,id):
    if request.method == 'POST':

        try:
            banner_image = request.FILES['bannerimage']
            banner_title = request.POST['bannertitle']
            banner_description = request.POST['bannerdescription']
            url =request.POST['url']
        except:
            messages.warning(request,'please fill  the columns')

            
        try:
                banner = Banner(
                    id = id,
                    banner_image = banner_image,
                    banner_title = banner_title,
                    banner_description = banner_description,
                    url = url
                )

                banner.save()
                return redirect('admin_banners')
            
        except Exception as e:
                print(e)
                messages.warning(request,'Something went Wrong')
                return redirect('admin_banners')

    return render(request,'banner.html')




@login_required(login_url='/admin')
@user_passes_test(admin_check)
def delete_banner(request,id):

    Banner.objects.filter(id = id).delete()
    messages.success(request,'Banner Deleted')
    return redirect('admin_banners')

    
# ==================================================================================Sales Report======================================================================================================
@method_decorator(never_cache, name='dispatch')
@method_decorator(login_required(login_url='/admin'), name='dispatch')
@method_decorator(user_passes_test(admin_check), name='dispatch')
class admin_Salesreport_view(TemplateView):
    template_name = 'salesreport.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        order = Order.objects.all()

        context['order'] = order

        return context
    


@method_decorator(login_required(login_url='/admin'), name='dispatch')
@method_decorator(user_passes_test(admin_check), name='dispatch')
class Admin_Salesreport_download(View):

    def get(self,request):
     
       try:
        start_str = self.request.GET.get('start')
        end_str = self.request.GET.get('end')
        start = datetime.strptime(start_str, '%Y-%m-%d').date()
        end = datetime.strptime(end_str, '%Y-%m-%d').date()
       except Exception as e:
          print(e)
          start = None
          
       if start:
           

           order = Order.objects.filter(Q(order_date__gte = start) & Q (order_date__lte = end))

           
       else:
           order = Order.objects.all()


       
       #pdf
       if 'excel' not in request.GET:


                buffer = BytesIO()
            
                # Create the PDF object
                pdf = SimpleDocTemplate(buffer, title  = 'Sales report')

                #  data for the table
                data = []
                header = ["Order Date", "Order ID", "Category", "Brand", "Sales Amount"]
                data.append(header)

                for report in order:
                    data.append([str(report.order_date.date()), str(report.id), report.product_id.category.category_name, report.product_id.brand.brand_name, str(report.amount)])

                # Creating  the table
                table = Table(data, colWidths=[100, 100, 100, 100, 100], rowHeights=30)

                table.setStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ])


                # page heading
                styles = getSampleStyleSheet()
                heading = Paragraph("Sales Report", style=styles["Heading1"])

                # writing the table in the pdf
                pdf.build([heading, table])

                # Close the PDF object
                buffer.seek(0)
                
                # Get the value of the PDF file from the buffer
                pdf = buffer.getvalue()

                # Return the PDF file through Django's FileResponse
                response = FileResponse(BytesIO(pdf), content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
                return response




            # excel
       else:

                # Creating a new Excel workbook
                    workbook = openpyxl.Workbook()

                    # Selecting the active worksheet
                    worksheet = workbook.active

                    # the headers to the worksheet
                    worksheet['A1'] = "Order Date"
                    worksheet['B1'] = "Order ID"
                    worksheet['D1'] = "Category"
                    worksheet['E1'] = "Brand"
                    worksheet['C1'] = "Sales Amount"


                    #  sales report data to the worksheet
                    for row, report in enumerate(order, start=2):
                        worksheet.cell(row=row, column=1, value= str(report.order_date.date()))
                        worksheet.cell(row=row, column=2, value= str(report.id))
                        worksheet.cell(row=row, column=4, value=report.product_id.category.category_name)
                        worksheet.cell(row=row, column=5, value=report.product_id.brand.brand_name)
                        worksheet.cell(row=row, column=3, value= str(report.amount))


                # Create a file-like buffer to receive Excel workbook data
                    buffer = io.BytesIO()

                    # Save the workbook to the buffer
                    workbook.save(buffer)

                    # FileResponse sets the Content-Disposition header so that browsers
                    # present the option to save the file.
                    buffer.seek(0)
                    return FileResponse(buffer, as_attachment=True,filename='sales_report.xlsx')





@method_decorator(login_required(login_url='/admin'), name='dispatch')
@method_decorator(user_passes_test(admin_check), name='dispatch')
class admin_salesreport_filter(View):

    def post(self,request):

        start_str = request.POST['from']
        end_str = request.POST['to']

        start =  datetime.strptime(start_str, '%Y-%m-%d').date()
        end =  datetime.strptime(end_str, '%Y-%m-%d').date()
        try:
          order = Order.objects.filter(Q(order_date__gte = start) & Q (order_date__lte = end))
          return render(request,'salesreport.html',{'order':order,'start':start_str,'end':end_str})
        except Exception as e:
            print(e)
            messages.warning(request,'Try with proper values')
            return redirect('admin_salesreport')
       
    







@never_cache
def admin_signin(request):
    if request.user.is_authenticated and request.user.is_superadmin ==True:
        return redirect('admin_pannel')
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']

        admin = auth.authenticate(email = email, password = password)

        if admin is not None and admin.is_superadmin:
            auth.login(request,admin)
            return redirect('admin_pannel')
        else:
            messages.warning(request,'Password or Email is invalid.Try again')
            return redirect('admin_signin')


    return render(request,'admin_signin.html')



@login_required(login_url='/admin')
@user_passes_test(admin_check)
def logout(request):
    auth.logout(request)
    return redirect('admin')




@method_decorator(never_cache, name='dispatch')
@method_decorator(login_required(login_url='/admin'), name='dispatch')
@method_decorator(user_passes_test(admin_check), name='dispatch')
class Admin_product(TemplateView):
    template_name = 'admin product.html'

    def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)

       product = Product.objects.all()

       context['product'] = product

       return context
    



@method_decorator(login_required(login_url='/admin'), name='dispatch')
@method_decorator(user_passes_test(admin_check), name='dispatch')
class Admin_product_details(TemplateView):
    template_name= 'admin_product_details'